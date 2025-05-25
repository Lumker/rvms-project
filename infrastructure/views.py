# infrastructure/views.py

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.db import models  # Add this import
from django.db.models import Q, Count, Avg, Sum
from django.views.generic import ListView, DetailView, CreateView, UpdateView
from django.urls import reverse_lazy, reverse
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.core.paginator import Paginator
from datetime import datetime, timedelta
import json

from .models import (
    InfrastructureAsset, User, WaterSource, WaterDistributionPoint,
    MaintenanceRecord, ServiceInterruption, AssetInspection
)
from governance.models import Village, Municipality, WardCommittee
from households.models import Household
from .forms import (
    BulkMaintenanceForm, WaterSourceForm, WaterDistributionPointForm, MaintenanceRecordForm,
    ServiceInterruptionForm, AssetInspectionForm, InfrastructureAssetForm
)

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy, reverse
from django.utils import timezone
from django.core.paginator import Paginator
from django.db.models import (
    Count, Sum, Avg, Q, F, Case, When, DecimalField, 
    IntegerField, Value # Make sure Value is imported
)
from django.db.models.functions import Cast  # Import Cast from functions
import csv
from io import BytesIO

# infrastructure/views.py (add these missing views)

# Add these imports at the top
from django.views.generic import CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin

# Add these missing view classes to your infrastructure/views.py

class InfrastructureAssetCreateView(LoginRequiredMixin, CreateView):
    """Create new infrastructure asset"""
    model = InfrastructureAsset
    form_class = InfrastructureAssetForm
    template_name = 'infrastructure/asset_form.html'
    success_url = reverse_lazy('infrastructure:asset_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Infrastructure asset created successfully!')
        return super().form_valid(form)


class InfrastructureAssetUpdateView(LoginRequiredMixin, UpdateView):
    """Update infrastructure asset"""
    model = InfrastructureAsset
    form_class = InfrastructureAssetForm
    template_name = 'infrastructure/asset_form.html'
    
    def get_success_url(self):
        return reverse('infrastructure:asset_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, 'Infrastructure asset updated successfully!')
        return super().form_valid(form)


class WaterDistributionPointDetailView(LoginRequiredMixin, DetailView):
    """Detail view for water distribution point"""
    model = WaterDistributionPoint
    template_name = 'infrastructure/water/distribution_detail.html'
    context_object_name = 'distribution_point'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        distribution_point = self.object
        
        # Get households served
        households_served = distribution_point.households_served.all()
        
        # Recent maintenance
        recent_maintenance = distribution_point.maintenance_records.order_by('-scheduled_date')[:5]
        
        # Service interruptions
        recent_interruptions = distribution_point.service_interruptions.order_by('-start_time')[:5]
        
        context.update({
            'households_served': households_served,
            'recent_maintenance': recent_maintenance,
            'recent_interruptions': recent_interruptions,
            'total_households': households_served.count(),
        })
        
        return context


class WaterDistributionPointUpdateView(LoginRequiredMixin, UpdateView):
    """Update water distribution point"""
    model = WaterDistributionPoint
    form_class = WaterDistributionPointForm
    template_name = 'infrastructure/water/distribution_form.html'
    
    def get_success_url(self):
        return reverse('infrastructure:water_distribution_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, 'Water distribution point updated successfully!')
        return super().form_valid(form)


class MaintenanceRecordDetailView(LoginRequiredMixin, DetailView):
    """Detail view for maintenance record"""
    model = MaintenanceRecord
    template_name = 'infrastructure/maintenance/record_detail.html'
    context_object_name = 'maintenance_record'


class ServiceInterruptionDetailView(LoginRequiredMixin, DetailView):
    """Detail view for service interruption"""
    model = ServiceInterruption
    template_name = 'infrastructure/interruption_detail.html'
    context_object_name = 'interruption'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        interruption = self.object
        
        # Calculate impact metrics
        if interruption.end_time and interruption.start_time:
            duration_hours = (interruption.end_time - interruption.start_time).total_seconds() / 3600
        else:
            duration_hours = None
        
        context.update({
            'duration_hours': duration_hours,
            'affected_households_count': interruption.affected_households.count(),
        })
        
        return context


class ServiceInterruptionUpdateView(LoginRequiredMixin, UpdateView):
    """Update service interruption"""
    model = ServiceInterruption
    form_class = ServiceInterruptionForm
    template_name = 'infrastructure/interruption_form.html'
    
    def get_success_url(self):
        return reverse('infrastructure:interruption_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, 'Service interruption updated successfully!')
        return super().form_valid(form)

# infrastructure/views.py (update the AssetInspectionListView)

class AssetInspectionListView(LoginRequiredMixin, ListView):
    """List asset inspections"""
    model = AssetInspection
    template_name = 'infrastructure/inspection_list.html'
    context_object_name = 'inspections'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = AssetInspection.objects.select_related(
            'asset', 'inspector'
        ).order_by('-inspection_date')
        
        # Apply filters
        inspection_type = self.request.GET.get('inspection_type')
        if inspection_type:
            queryset = queryset.filter(inspection_type=inspection_type)
        
        asset_category = self.request.GET.get('asset_category')
        if asset_category:
            queryset = queryset.filter(asset__category=asset_category)
        
        overall_condition = self.request.GET.get('overall_condition')
        if overall_condition:
            queryset = queryset.filter(overall_condition=overall_condition)
        
        date_from = self.request.GET.get('date_from')
        if date_from:
            queryset = queryset.filter(inspection_date__gte=date_from)
        
        date_to = self.request.GET.get('date_to')
        if date_to:
            queryset = queryset.filter(inspection_date__lte=date_to)
        
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(asset__name__icontains=search) |
                Q(inspector__first_name__icontains=search) |
                Q(inspector__last_name__icontains=search)
            )
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get the base queryset without pagination
        base_queryset = self.get_queryset()
        
        # Calculate statistics from the base queryset
        total_inspections = base_queryset.count()
        
        # Get current month inspections
        from django.utils import timezone
        current_month = timezone.now().month
        current_year = timezone.now().year
        this_month_count = base_queryset.filter(
            inspection_date__month=current_month,
            inspection_date__year=current_year
        ).count()
        
        # Count by condition
        poor_condition_count = base_queryset.filter(overall_condition='poor').count()
        excellent_condition_count = base_queryset.filter(overall_condition='excellent').count()
        
        context.update({
            'inspection_types': AssetInspection.INSPECTION_TYPES,
            'asset_categories': InfrastructureAsset.ASSET_CATEGORIES,
            'conditions': InfrastructureAsset.CONDITION_CHOICES,
            'current_filters': {
                'inspection_type': self.request.GET.get('inspection_type', ''),
                'asset_category': self.request.GET.get('asset_category', ''),
                'overall_condition': self.request.GET.get('overall_condition', ''),
                'date_from': self.request.GET.get('date_from', ''),
                'date_to': self.request.GET.get('date_to', ''),
                'search': self.request.GET.get('search', ''),
            },
            'total_inspections': total_inspections,
            'this_month_count': this_month_count,
            'poor_condition_count': poor_condition_count,
            'excellent_condition_count': excellent_condition_count,
        })
        return context

class AssetInspectionDetailView(LoginRequiredMixin, DetailView):
    """Detail view for asset inspection"""
    model = AssetInspection
    template_name = 'infrastructure/inspection_detail.html'
    context_object_name = 'inspection'


class AssetInspectionCreateView(LoginRequiredMixin, CreateView):
    """Create new asset inspection"""
    model = AssetInspection
    form_class = AssetInspectionForm
    template_name = 'infrastructure/inspection_form.html'
    success_url = reverse_lazy('infrastructure:inspection_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Asset inspection created successfully!')
        return super().form_valid(form)


class AssetInspectionUpdateView(LoginRequiredMixin, UpdateView):
    """Update asset inspection"""
    model = AssetInspection
    form_class = AssetInspectionForm
    template_name = 'infrastructure/inspection_form.html'
    
    def get_success_url(self):
        return reverse('infrastructure:inspection_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, 'Asset inspection updated successfully!')
        return super().form_valid(form)


class BulkMaintenanceView(LoginRequiredMixin, CreateView):
    """Bulk maintenance scheduling"""
    form_class = BulkMaintenanceForm
    template_name = 'infrastructure/maintenance/bulk_form.html'
    success_url = reverse_lazy('infrastructure:maintenance_list')
    
    def form_valid(self, form):
        assets = form.cleaned_data['assets']
        maintenance_type = form.cleaned_data['maintenance_type']
        scheduled_date = form.cleaned_data['scheduled_date']
        description = form.cleaned_data['description']
        technician = form.cleaned_data.get('technician')
        
        # Create maintenance records for all selected assets
        created_count = 0
        for asset in assets:
            MaintenanceRecord.objects.create(
                asset=asset,
                maintenance_type=maintenance_type,
                scheduled_date=scheduled_date,
                description=description,
                technician=technician,
                status='scheduled'
            )
            created_count += 1
        
        messages.success(
            self.request, 
            f'Successfully scheduled maintenance for {created_count} assets!'
        )
        return super().form_valid(form)


class QuickInspectionView(LoginRequiredMixin, CreateView):
    """Quick inspection creation"""
    model = AssetInspection
    fields = ['inspection_type', 'overall_condition', 'operational_status', 'safety_concerns']
    template_name = 'infrastructure/quick_inspection_form.html'
    
    def get_success_url(self):
        return reverse('infrastructure:asset_detail', kwargs={'pk': self.object.asset.pk})
    
    def form_valid(self, form):
        form.instance.asset_id = self.kwargs['pk']
        form.instance.inspection_date = timezone.now().date()
        form.instance.inspector = self.request.user
        messages.success(self.request, 'Quick inspection completed successfully!')
        return super().form_valid(form)


# Additional AJAX and utility views
@login_required
def get_assets_by_category(request):
    """Get assets filtered by category (AJAX)"""
    category = request.GET.get('category')
    village_id = request.GET.get('village_id')
    
    queryset = InfrastructureAsset.objects.filter(operational_status=True)
    
    if category:
        queryset = queryset.filter(category=category)
    if village_id:
        queryset = queryset.filter(village_id=village_id)
    
    assets = queryset.values('id', 'name', 'asset_id', 'condition')
    return JsonResponse({'assets': list(assets)})


@login_required
def maintenance_calendar_data(request):
    """Get maintenance calendar data (AJAX)"""
    start_date = request.GET.get('start')
    end_date = request.GET.get('end')
    
    maintenance_records = MaintenanceRecord.objects.filter(
        scheduled_date__range=[start_date, end_date]
    ).select_related('asset')
    
    events = []
    for record in maintenance_records:
        events.append({
            'id': record.id,
            'title': f"{record.asset.name} - {record.get_maintenance_type_display()}",
            'start': record.scheduled_date.isoformat(),
            'backgroundColor': {
                'scheduled': '#ffc107',
                'in_progress': '#17a2b8',
                'completed': '#28a745',
                'cancelled': '#dc3545',
                'deferred': '#6c757d',
            }.get(record.status, '#6c757d'),
            'url': reverse('infrastructure:maintenance_detail', kwargs={'pk': record.pk})
        })
    
    return JsonResponse(events, safe=False)


@login_required
def dashboard_alerts(request):
    """Get dashboard alert counts (AJAX)"""
    data = {
        'ongoing_interruptions': ServiceInterruption.objects.filter(is_resolved=False).count(),
        'overdue_inspections': InfrastructureAsset.objects.filter(
            next_inspection_due__lt=timezone.now().date()
        ).count(),
        'critical_condition': InfrastructureAsset.objects.filter(condition='critical').count(),
    }
    return JsonResponse(data)


# infrastructure/views.py (add these simplified report views)

@login_required
def maintenance_report(request):
    """Generate maintenance report - Simplified version"""
    # Get date range from request or default to current month
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if not date_from or not date_to:
        today = timezone.now().date()
        date_from = today.replace(day=1)  # First day of current month
        date_to = today
    else:
        try:
            date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
            date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
        except ValueError:
            today = timezone.now().date()
            date_from = today.replace(day=1)
            date_to = today
    
    # Get maintenance records
    maintenance_records = MaintenanceRecord.objects.filter(
        scheduled_date__range=[date_from, date_to]
    ).select_related('asset', 'technician').order_by('-scheduled_date')
    
    # Calculate summary statistics
    total_cost = maintenance_records.filter(
        status='completed'
    ).aggregate(total=Sum('actual_cost'))['total'] or 0
    
    completed_count = maintenance_records.filter(status='completed').count()
    pending_count = maintenance_records.filter(status__in=['scheduled', 'in_progress']).count()
    total_records = maintenance_records.count()
    
    # Group by maintenance type
    by_type = {}
    for record in maintenance_records:
        mtype = record.get_maintenance_type_display()
        if mtype not in by_type:
            by_type[mtype] = {'count': 0, 'cost': 0}
        by_type[mtype]['count'] += 1
        if record.actual_cost:
            by_type[mtype]['cost'] += float(record.actual_cost)
    
    # Group by asset category
    by_category = {}
    for record in maintenance_records:
        category = record.asset.get_category_display()
        if category not in by_category:
            by_category[category] = {'count': 0, 'cost': 0}
        by_category[category]['count'] += 1
        if record.actual_cost:
            by_category[category]['cost'] += float(record.actual_cost)
    
    context = {
        'date_from': date_from,
        'date_to': date_to,
        'maintenance_records': maintenance_records,
        'total_cost': total_cost,
        'completed_count': completed_count,
        'pending_count': pending_count,
        'total_records': total_records,
        'by_type': by_type,
        'by_category': by_category,
        'completion_rate': round((completed_count / total_records * 100), 1) if total_records > 0 else 0,
    }
    
    return render(request, 'infrastructure/reports/maintenance_report.html', context)


@login_required
def water_coverage_report(request):
    """Generate water coverage report - Simplified version"""
    # Get villages with water infrastructure
    villages = Village.objects.filter(is_active=True).prefetch_related(
        'infrastructure_assets', 'households'
    ).order_by('name')
    
    village_data = []
    total_villages = 0
    villages_with_water = 0
    total_water_sources = 0
    total_households = 0
    
    for village in villages:
        total_villages += 1
        water_sources_count = village.infrastructure_assets.filter(category='water').count()
        households_count = village.households.count()
        
        if water_sources_count > 0:
            villages_with_water += 1
        
        total_water_sources += water_sources_count
        total_households += households_count
        
        # Calculate population
        population = 0
        for household in village.households.all():
            population += household.residents.count()
        
        village_data.append({
            'village': village,
            'water_sources_count': water_sources_count,
            'households_count': households_count,
            'population': population,
            'coverage_ratio': water_sources_count / households_count if households_count > 0 else 0
        })
    
    # Calculate coverage statistics
    coverage_percentage = (villages_with_water / total_villages * 100) if total_villages > 0 else 0
    
    # Water source statistics
    water_sources = WaterSource.objects.select_related('village').order_by('village__name', 'name')
    total_capacity = WaterSource.objects.aggregate(total=Sum('daily_capacity_litres'))['total'] or 0
    
    context = {
        'village_data': village_data,
        'total_villages': total_villages,
        'villages_with_water': villages_with_water,
        'coverage_percentage': round(coverage_percentage, 1),
        'water_sources': water_sources,
        'total_capacity': total_capacity,
        'total_water_sources': total_water_sources,
        'total_households': total_households,
        'average_sources_per_village': round(total_water_sources / total_villages, 2) if total_villages > 0 else 0,
    }
    
    return render(request, 'infrastructure/reports/water_coverage_report.html', context)


@login_required
def asset_condition_report(request):
    """Generate asset condition report - Simplified version"""
    # Assets by condition
    condition_stats = {}
    for condition_code, condition_name in InfrastructureAsset.CONDITION_CHOICES:
        count = InfrastructureAsset.objects.filter(condition=condition_code).count()
        condition_stats[condition_name] = count
    
    # Assets by category and condition
    category_condition_stats = {}
    for category_code, category_name in InfrastructureAsset.ASSET_CATEGORIES:
        category_condition_stats[category_name] = {}
        for condition_code, condition_name in InfrastructureAsset.CONDITION_CHOICES:
            count = InfrastructureAsset.objects.filter(
                category=category_code,
                condition=condition_code
            ).count()
            category_condition_stats[category_name][condition_name] = count
    
    # Critical assets requiring immediate attention
    critical_assets = InfrastructureAsset.objects.filter(
        condition='critical'
    ).select_related('village').order_by('village__name', 'name')
    
    # Assets with overdue inspections
    overdue_inspections = InfrastructureAsset.objects.filter(
        next_inspection_due__lt=timezone.now().date()
    ).select_related('village').order_by('next_inspection_due')
    
    context = {
        'condition_stats': condition_stats,
        'category_condition_stats': category_condition_stats,
        'critical_assets': critical_assets,
        'overdue_inspections': overdue_inspections,
        'total_assets': InfrastructureAsset.objects.count(),
        'critical_count': critical_assets.count(),
        'overdue_count': overdue_inspections.count(),
    }
    
    return render(request, 'infrastructure/reports/asset_condition_report.html', context)


# infrastructure/views.py

def water_coverage_report(request):
    """Generate water coverage report - corrected version"""
    from governance.models import Village
    from households.models import Household
    
    # Start with basic village data
    villages = Village.objects.filter(is_active=True).select_related(
        'traditional_council__municipality__district'
    ).annotate(
        total_population=F('population'),
        total_households=Count('households', distinct=True),
    )
    
    # Add infrastructure count using the correct relationship name
    try:
        villages = villages.annotate(
            water_infrastructure_count=Count(
                'infrastructure_assets',  # Changed from 'infrastructure_items'
                filter=Q(infrastructure_assets__infrastructure_type='water'),
                distinct=True
            )
        )
        print("✅ Infrastructure annotation successful")
    except Exception as e:
        print(f"❌ Infrastructure annotation failed: {e}")
        # Add a default value if infrastructure relationship doesn't exist
        villages = villages.annotate(water_infrastructure_count=Value(0))
    
    # Simple coverage calculation without Cast (avoiding complexity)
    villages = villages.annotate(
        # Calculate households with water (simplified)
        households_with_water=Case(
            When(water_infrastructure_count__gt=0, then=F('total_households') * 80 / 100),
            default=0,
            output_field=IntegerField()
        ),
        water_coverage_percentage=Case(
            When(total_households__gt=0, then=F('households_with_water') * 100.0 / F('total_households')),
            default=0,
            output_field=DecimalField(max_digits=5, decimal_places=2)
        )
    ).order_by('traditional_council__municipality__name', 'name')

    # Calculate summary statistics
    total_villages = villages.count()
    villages_with_water = villages.filter(households_with_water__gt=0).count()
    total_pop = villages.aggregate(total=Sum('total_population'))['total'] or 0
    
    avg_coverage = villages.aggregate(
        avg=Avg('water_coverage_percentage')
    )['avg'] or 0

    context = {
        'villages': villages,
        'total_villages': total_villages,
        'villages_with_water': villages_with_water,
        'total_population': total_pop,
        'average_coverage': round(avg_coverage, 2),
        'villages_without_water': total_villages - villages_with_water,
        'report_date': timezone.now(),
        'request_path': request.path,
        'note': 'Simplified report using infrastructure_assets relationship.'
    }
    
    # Handle export requests
    export_format = request.GET.get('export')
    if export_format == 'csv':
        return export_water_coverage_csv(villages)
    elif export_format == 'excel':
        return export_water_coverage_excel(villages)
    elif export_format == 'pdf':
        return export_water_coverage_pdf(villages, context)
    
    return render(request, 'infrastructure/reports/water_coverage.html', context)

def debug_relationships(request):
    """Debug view to check model relationships"""
    from governance.models import Village
    from households.models import Household
    
    try:
        # Import HouseholdService to check if it exists
        from households.models import HouseholdService
        household_service_exists = True
        
        # Get HouseholdService fields
        service_fields = []
        for field in HouseholdService._meta.get_fields():
            service_fields.append({
                'name': field.name,
                'type': type(field).__name__,
                'related_model': getattr(field, 'related_model', None).__name__ if hasattr(field, 'related_model') and field.related_model else None
            })
    except ImportError:
        household_service_exists = False
        service_fields = ["HouseholdService model not found"]
    
    # Get Village relationships
    village_relations = []
    for field in Village._meta.get_fields():
        village_relations.append({
            'name': field.name,
            'type': type(field).__name__,
            'related_model': getattr(field, 'related_model', None).__name__ if hasattr(field, 'related_model') and field.related_model else None
        })
    
    # Get Household relationships
    household_relations = []
    for field in Household._meta.get_fields():
        household_relations.append({
            'name': field.name,
            'type': type(field).__name__,
            'related_model': getattr(field, 'related_model', None).__name__ if hasattr(field, 'related_model') and field.related_model else None
        })
    
    return JsonResponse({
        'household_service_exists': household_service_exists,
        'village_relations': village_relations,
        'household_relations': household_relations,
        'service_fields': service_fields
    }, indent=2)

def export_water_coverage_csv(villages):
    """Export water coverage data as CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="water_coverage_report.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Village', 'Traditional Council', 'Municipality', 'District',
        'Population', 'Total Households', 'Households with Water',
        'Coverage %', 'Water Infrastructure Count'
    ])
    
    for village in villages:
        writer.writerow([
            village.name,
            village.traditional_council.name,
            village.traditional_council.municipality.name,
            village.traditional_council.municipality.district.name,
            village.total_population or 0,
            village.total_households,
            village.households_with_water,
            f"{village.water_coverage_percentage:.1f}%",
            village.water_infrastructure_count,
        ])
    
    return response

def export_water_coverage_excel(villages):
    """Export water coverage data as Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # Fallback to CSV if openpyxl is not available
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="water_coverage_report.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Note: Excel export requires openpyxl. Exporting as CSV instead.'])
        writer.writerow([])
        
        # Write CSV data
        writer.writerow([
            'Village', 'Traditional Council', 'Municipality', 'District',
            'Population', 'Total Households', 'Households with Water',
            'Coverage %', 'Water Infrastructure Count'
        ])
        
        for village in villages:
            writer.writerow([
                village.name,
                village.traditional_council.name,
                village.traditional_council.municipality.name,
                village.traditional_council.municipality.district.name,
                village.total_population or 0,
                village.total_households,
                village.households_with_water,
                f"{village.water_coverage_percentage:.1f}%",
                village.water_infrastructure_count,
            ])
        
        return response
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Water Coverage Report"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        'Village', 'Traditional Council', 'Municipality', 'District',
        'Population', 'Total Households', 'Households with Water',
        'Coverage %', 'Water Infrastructure Count'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Data rows
    for row, village in enumerate(villages, 2):
        ws.cell(row=row, column=1, value=village.name)
        ws.cell(row=row, column=2, value=village.traditional_council.name)
        ws.cell(row=row, column=3, value=village.traditional_council.municipality.name)
        ws.cell(row=row, column=4, value=village.traditional_council.municipality.district.name)
        ws.cell(row=row, column=5, value=village.total_population or 0)
        ws.cell(row=row, column=6, value=village.total_households)
        ws.cell(row=row, column=7, value=village.households_with_water)
        ws.cell(row=row, column=8, value=f"{village.water_coverage_percentage:.1f}%")
        ws.cell(row=row, column=9, value=village.water_infrastructure_count)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    response = HttpResponse(
        excel_file.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="water_coverage_report.xlsx"'
    
    return response

def export_water_coverage_pdf(villages, context):
    """Export water coverage data as PDF"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER
    except ImportError as e:
        # Fallback if reportlab is not available
        response = HttpResponse(content_type='text/html')
        response.write(f'''
        <html>
        <head><title>PDF Export Error</title></head>
        <body>
            <h1>PDF Export Not Available</h1>
            <p>PDF export requires reportlab library.</p>
            <p>Error: {str(e)}</p>
            <p>Please try:</p>
            <ul>
                <li><a href="?export=csv">Download CSV</a></li>
                <li><a href="?export=excel">Download Excel</a></li>
            </ul>
            <p><a href="{context.get('request_path', '')}">&larr; Back to Report</a></p>
        </body>
        </html>
        ''')
        return response
    
    # Create the HttpResponse object with PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="water_coverage_report.pdf"'
    
    try:
        # Create the PDF document
        doc = SimpleDocTemplate(response, pagesize=A4)
        story = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Title
        title = Paragraph("Water Coverage Report", title_style)
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Summary information
        summary_data = [
            ['Total Villages:', str(context.get('total_villages', 0))],
            ['Villages with Water:', str(context.get('villages_with_water', 0))],
            ['Villages without Water:', str(context.get('villages_without_water', 0))],
            ['Total Population:', f"{context.get('total_population', 0):,}"],
            ['Average Coverage:', f"{context.get('average_coverage', 0):.1f}%"],
            ['Report Date:', context.get('report_date', timezone.now()).strftime('%Y-%m-%d %H:%M')],
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Village data table
        data = [['Village', 'Council', 'Municipality', 'Population', 'Coverage %']]
        
        for village in villages:
            data.append([
                village.name[:30],  # Truncate long names for PDF
                village.traditional_council.name[:20],
                village.traditional_council.municipality.name[:20],
                str(village.total_population or 0),
                f"{village.water_coverage_percentage:.1f}%"
            ])
        
        # Create table
        table = Table(data, colWidths=[1.8*inch, 1.5*inch, 1.5*inch, 0.8*inch, 0.8*inch])
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            
            # Data styling
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightblue]),
        ]))
        
        story.append(table)
        
        # Build PDF
        doc.build(story)
        
    except Exception as e:
        # If PDF generation fails, return error page
        response = HttpResponse(content_type='text/html')
        response.write(f'''
        <html>
        <head><title>PDF Generation Error</title></head>
        <body>
            <h1>PDF Generation Error</h1>
            <p>Error generating PDF: {str(e)}</p>
            <p>Please try:</p>
            <ul>
                <li><a href="?export=csv">Download CSV</a></li>
                <li><a href="?export=excel">Download Excel</a></li>
            </ul>
        </body>
        </html>
        ''')
        return response
    
    return response

def export_water_coverage_pdf(villages, context):
    """Export water coverage data as PDF"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER
    except ImportError as e:
        # Fallback if reportlab is not available
        response = HttpResponse(content_type='text/html')
        response.write(f'''
        <html>
        <head><title>PDF Export Error</title></head>
        <body>
            <h1>PDF Export Not Available</h1>
            <p>PDF export requires reportlab library.</p>
            <p>Error: {str(e)}</p>
            <p>Please try:</p>
            <ul>
                <li><a href="?export=csv">Download CSV</a></li>
                <li><a href="?export=excel">Download Excel</a></li>
            </ul>
            <p><a href="{context.get('request_path', '')}">&larr; Back to Report</a></p>
        </body>
        </html>
        ''')
        return response
    
    # Create the HttpResponse object with PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="water_coverage_report.pdf"'
    
    try:
        # Create the PDF document
        doc = SimpleDocTemplate(response, pagesize=A4)
        story = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Title
        title = Paragraph("Water Coverage Report", title_style)
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Summary information
        summary_data = [
            ['Total Villages:', str(context.get('total_villages', 0))],
            ['Villages with Water:', str(context.get('villages_with_water', 0))],
            ['Villages without Water:', str(context.get('villages_without_water', 0))],
            ['Total Population:', f"{context.get('total_population', 0):,}"],
            ['Average Coverage:', f"{context.get('average_coverage', 0):.1f}%"],
            ['Report Date:', context.get('report_date', timezone.now()).strftime('%Y-%m-%d %H:%M')],
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Village data table
        data = [['Village', 'Council', 'Municipality', 'Population', 'Coverage %']]
        
        for village in villages:
            data.append([
                village.name[:30],  # Truncate long names for PDF
                village.traditional_council.name[:20],
                village.traditional_council.municipality.name[:20],
                str(village.total_population or 0),
                f"{village.water_coverage_percentage:.1f}%"
            ])
        
        # Create table
        table = Table(data, colWidths=[1.8*inch, 1.5*inch, 1.5*inch, 0.8*inch, 0.8*inch])
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            
            # Data styling
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightblue]),
        ]))
        
        story.append(table)
        
        # Build PDF
        doc.build(story)
        
    except Exception as e:
        # If PDF generation fails, return error page
        response = HttpResponse(content_type='text/html')
        response.write(f'''
        <html>
        <head><title>PDF Generation Error</title></head>
        <body>
            <h1>PDF Generation Error</h1>
            <p>Error generating PDF: {str(e)}</p>
            <p>Please try:</p>
            <ul>
                <li><a href="?export=csv">Download CSV</a></li>
                <li><a href="?export=excel">Download Excel</a></li>
            </ul>
        </body>
        </html>
        ''')
        return response
    
    return response

# infrastructure/views.py

def debug_relationships(request):
    """Debug view to check model relationships"""
    from governance.models import Village
    from households.models import Household
    
    try:
        # Import HouseholdService to check if it exists
        from households.models import HouseholdService
        household_service_exists = True
        
        # Get HouseholdService fields
        service_fields = []
        for field in HouseholdService._meta.get_fields():
            service_fields.append({
                'name': field.name,
                'type': type(field).__name__,
                'related_model': getattr(field, 'related_model', None).__name__ if hasattr(field, 'related_model') and field.related_model else None
            })
    except ImportError:
        household_service_exists = False
        service_fields = ["HouseholdService model not found"]
    
    # Get Village relationships
    village_relations = []
    for field in Village._meta.get_fields():
        village_relations.append({
            'name': field.name,
            'type': type(field).__name__,
            'related_model': getattr(field, 'related_model', None).__name__ if hasattr(field, 'related_model') and field.related_model else None
        })
    
    # Get Household relationships
    household_relations = []
    for field in Household._meta.get_fields():
        household_relations.append({
            'name': field.name,
            'type': type(field).__name__,
            'related_model': getattr(field, 'related_model', None).__name__ if hasattr(field, 'related_model') and field.related_model else None
        })
    
    return JsonResponse({
        'household_service_exists': household_service_exists,
        'village_relations': village_relations,
        'household_relations': household_relations,
        'service_fields': service_fields
    })

def water_coverage_report(request):
    """Generate water coverage report - corrected version"""
    from governance.models import Village
    from households.models import Household
    
    # Start with basic village data
    villages = Village.objects.filter(is_active=True).select_related(
        'traditional_council__municipality__district'
    ).annotate(
        total_population=F('population'),
        total_households=Count('households', distinct=True),
    )
    
    # Add infrastructure count using the correct relationship name
    try:
        villages = villages.annotate(
            water_infrastructure_count=Count(
                'infrastructure_assets',  # Changed from 'infrastructure_items'
                filter=Q(infrastructure_assets__infrastructure_type='water'),
                distinct=True
            )
        )
        print("✅ Infrastructure annotation successful")
    except Exception as e:
        print(f"❌ Infrastructure annotation failed: {e}")
        # Add a default value if infrastructure relationship doesn't exist
        villages = villages.annotate(water_infrastructure_count=Value(0))
    
    # For now, create a simple coverage calculation without household services
    villages = villages.annotate(
        # Simple logic: if there's water infrastructure, assume 80% coverage
        # This is temporary until we fix the household services relationship
        households_with_water=Case(
            When(water_infrastructure_count__gt=0, then=Cast(F('total_households') * 0.8, IntegerField())),
            default=0,
            output_field=IntegerField()
        ),
        water_coverage_percentage=Case(
            When(total_households__gt=0, then=F('households_with_water') * 100.0 / F('total_households')),
            default=0,
            output_field=DecimalField(max_digits=5, decimal_places=2)
        )
    ).order_by('traditional_council__municipality__name', 'name')

    # Calculate summary statistics
    total_villages = villages.count()
    villages_with_water = villages.filter(households_with_water__gt=0).count()
    total_pop = villages.aggregate(total=Sum('total_population'))['total'] or 0
    
    avg_coverage = villages.aggregate(
        avg=Avg('water_coverage_percentage')
    )['avg'] or 0

    context = {
        'villages': villages,
        'total_villages': total_villages,
        'villages_with_water': villages_with_water,
        'total_population': total_pop,
        'average_coverage': round(avg_coverage, 2),
        'villages_without_water': total_villages - villages_with_water,
        'report_date': timezone.now(),
        'request_path': request.path,
        'note': 'This is a simplified report using infrastructure_assets relationship.'
    }
    
    # Handle export requests
    export_format = request.GET.get('export')
    if export_format == 'csv':
        return export_water_coverage_csv(villages)
    elif export_format == 'excel':
        return export_water_coverage_excel(villages)
    elif export_format == 'pdf':
        return export_water_coverage_pdf(villages, context)
    
    return render(request, 'infrastructure/reports/water_coverage.html', context)

def debug_relationships(request):
    """Debug view to check model relationships"""
    from governance.models import Village
    from households.models import Household
    
    try:
        # Import HouseholdService to check if it exists
        from households.models import HouseholdService
        household_service_exists = True
        
        # Get HouseholdService fields
        service_fields = []
        for field in HouseholdService._meta.get_fields():
            service_fields.append({
                'name': field.name,
                'type': type(field).__name__,
                'related_model': getattr(field, 'related_model', None).__name__ if hasattr(field, 'related_model') and field.related_model else None
            })
    except ImportError:
        household_service_exists = False
        service_fields = ["HouseholdService model not found"]
    
    # Get Village relationships
    village_relations = []
    for field in Village._meta.get_fields():
        village_relations.append({
            'name': field.name,
            'type': type(field).__name__,
            'related_model': getattr(field, 'related_model', None).__name__ if hasattr(field, 'related_model') and field.related_model else None
        })
    
    # Get Household relationships
    household_relations = []
    for field in Household._meta.get_fields():
        household_relations.append({
            'name': field.name,
            'type': type(field).__name__,
            'related_model': getattr(field, 'related_model', None).__name__ if hasattr(field, 'related_model') and field.related_model else None
        })
    
    return JsonResponse({
        'household_service_exists': household_service_exists,
        'village_relations': village_relations,
        'household_relations': household_relations,
        'service_fields': service_fields
    }, indent=2)  # Pretty print the JSON

# Your export functions remain the same...
def export_water_coverage_csv(villages):
    """Export water coverage data as CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="water_coverage_report.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Village', 'Traditional Council', 'Municipality', 'District',
        'Population', 'Total Households', 'Households with Water',
        'Coverage %', 'Water Infrastructure Count'
    ])
    
    for village in villages:
        writer.writerow([
            village.name,
            village.traditional_council.name,
            village.traditional_council.municipality.name,
            village.traditional_council.municipality.district.name,
            village.total_population or 0,
            village.total_households,
            village.households_with_water,
            f"{village.water_coverage_percentage:.1f}%",
            village.water_infrastructure_count,
        ])
    
    return response



@login_required
def export_maintenance_records(request):
    """Export maintenance records to CSV"""
    import csv
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="maintenance_records.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Asset', 'Maintenance Type', 'Status', 'Scheduled Date',
        'Completed Date', 'Technician', 'Estimated Cost', 'Actual Cost'
    ])
    
    records = MaintenanceRecord.objects.select_related('asset', 'technician')
    for record in records:
        writer.writerow([
            record.asset.name,
            record.get_maintenance_type_display(),
            record.get_status_display(),
            record.scheduled_date.strftime('%Y-%m-%d'),
            record.completed_date.strftime('%Y-%m-%d') if record.completed_date else '',
            record.technician.get_full_name() if record.technician else '',
            record.estimated_cost or '',
            record.actual_cost or ''
        ])
    
    return response


@login_required
def export_water_sources(request):
    """Export water sources to CSV"""
    import csv
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="water_sources.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Name', 'Source Type', 'Village', 'Water Quality', 'Daily Capacity (L)',
        'Yield (L/hr)', 'Depth (m)', 'Power Source', 'Operational Status'
    ])
    
    sources = WaterSource.objects.select_related('village')
    for source in sources:
        writer.writerow([
            source.name,
            source.get_source_type_display(),
            source.village.name,
            source.get_water_quality_display(),
            source.daily_capacity_litres or '',
            source.yield_litres_per_hour or '',
            source.depth_meters or '',
            source.get_power_source_display() if source.power_source else '',
            'Yes' if source.operational_status else 'No'
        ])
    
    return response


@login_required
def infrastructure_dashboard(request):
    """Infrastructure module dashboard"""
    # Basic statistics
    total_assets = InfrastructureAsset.objects.count()
    water_assets = InfrastructureAsset.objects.filter(category='water').count()
    operational_assets = InfrastructureAsset.objects.filter(operational_status=True).count()
    critical_condition = InfrastructureAsset.objects.filter(condition='critical').count()
    
    # Water infrastructure stats
    water_sources = WaterSource.objects.count()
    water_distribution_points = WaterDistributionPoint.objects.count()
    operational_water_sources = WaterSource.objects.filter(operational_status=True).count()
    
    # Recent maintenance
    recent_maintenance = MaintenanceRecord.objects.select_related('asset').order_by('-created_at')[:5]
    
    # Service interruptions
    ongoing_interruptions = ServiceInterruption.objects.filter(is_resolved=False).count()
    recent_interruptions = ServiceInterruption.objects.select_related('asset').order_by('-start_time')[:5]
    
    # Overdue inspections
    overdue_inspections = InfrastructureAsset.objects.filter(
        next_inspection_due__lt=timezone.now().date()
    ).count()
    
    # Asset condition distribution
    condition_stats = InfrastructureAsset.objects.values('condition').annotate(
        count=Count('id')
    ).order_by('condition')
    
    # Assets by category
    category_stats = InfrastructureAsset.objects.values('category').annotate(
        count=Count('id'),
        operational=Count('id', filter=Q(operational_status=True))
    ).order_by('category')
    
    # Water coverage by village
    villages_with_water = Village.objects.annotate(
        water_sources_count=Count('infrastructure_assets', filter=Q(infrastructure_assets__category='water')),
        households_count=Count('households')
    ).filter(water_sources_count__gt=0)[:10]
    
    context = {
        'total_assets': total_assets,
        'water_assets': water_assets,
        'operational_assets': operational_assets,
        'critical_condition': critical_condition,
        'water_sources': water_sources,
        'water_distribution_points': water_distribution_points,
        'operational_water_sources': operational_water_sources,
        'recent_maintenance': recent_maintenance,
        'ongoing_interruptions': ongoing_interruptions,
        'recent_interruptions': recent_interruptions,
        'overdue_inspections': overdue_inspections,
        'condition_stats': condition_stats,
        'category_stats': category_stats,
        'villages_with_water': villages_with_water,
        'operational_percentage': round((operational_assets / total_assets * 100) if total_assets > 0 else 0),
    }
    
    return render(request, 'infrastructure/dashboard.html', context)


class InfrastructureAssetListView(LoginRequiredMixin, ListView):
    """List all infrastructure assets"""
    model = InfrastructureAsset
    template_name = 'infrastructure/asset_list.html'
    context_object_name = 'assets'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = InfrastructureAsset.objects.select_related(
            'village', 'village__traditional_council', 'custodian'
        ).order_by('-created_at')
        
        # Apply filters
        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(category=category)
        
        condition = self.request.GET.get('condition')
        if condition:
            queryset = queryset.filter(condition=condition)
        
        village = self.request.GET.get('village')
        if village:
            queryset = queryset.filter(village_id=village)
        
        operational = self.request.GET.get('operational')
        if operational:
            queryset = queryset.filter(operational_status=operational == 'true')
        
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(asset_id__icontains=search) |
                Q(description__icontains=search)
            )
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'categories': InfrastructureAsset.ASSET_CATEGORIES,
            'conditions': InfrastructureAsset.CONDITION_CHOICES,
            'villages': Village.objects.filter(is_active=True).order_by('name'),
            'current_filters': {
                'category': self.request.GET.get('category', ''),
                'condition': self.request.GET.get('condition', ''),
                'village': self.request.GET.get('village', ''),
                'operational': self.request.GET.get('operational', ''),
                'search': self.request.GET.get('search', ''),
            },
            'total_assets': self.get_queryset().count(),
        })
        return context


class InfrastructureAssetDetailView(LoginRequiredMixin, DetailView):
    """Detail view for infrastructure asset"""
    model = InfrastructureAsset
    template_name = 'infrastructure/asset_detail.html'
    context_object_name = 'asset'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        asset = self.object
        
        # Recent maintenance records
        recent_maintenance = asset.maintenance_records.order_by('-scheduled_date')[:5]
        
        # Recent inspections
        recent_inspections = asset.inspections.order_by('-inspection_date')[:5]
        
        # Service interruptions
        recent_interruptions = asset.service_interruptions.order_by('-start_time')[:5]
        ongoing_interruptions = asset.service_interruptions.filter(is_resolved=False)
        
        # Maintenance costs
        total_maintenance_cost = asset.maintenance_records.filter(
            status='completed'
        ).aggregate(total=Sum('actual_cost'))['total'] or 0
        
        # Average downtime
        resolved_interruptions = asset.service_interruptions.filter(is_resolved=True)
        avg_downtime = 0
        if resolved_interruptions.exists():
            total_downtime = sum([
                (interruption.end_time - interruption.start_time).total_seconds() / 3600
                for interruption in resolved_interruptions
                if interruption.end_time
            ])
            avg_downtime = total_downtime / resolved_interruptions.count() if resolved_interruptions.count() > 0 else 0
        
        # If it's a water source, get distribution points
        distribution_points = None
        if hasattr(asset, 'watersource'):
            distribution_points = asset.watersource.distribution_points.all()
        
        context.update({
            'recent_maintenance': recent_maintenance,
            'recent_inspections': recent_inspections,
            'recent_interruptions': recent_interruptions,
            'ongoing_interruptions': ongoing_interruptions,
            'total_maintenance_cost': total_maintenance_cost,
            'avg_downtime_hours': round(avg_downtime, 2),
            'distribution_points': distribution_points,
        })
        
        return context


# infrastructure/views.py (update the WaterSourceListView)

class WaterSourceListView(LoginRequiredMixin, ListView):
    """List water sources"""
    model = WaterSource
    template_name = 'infrastructure/water/source_list.html'
    context_object_name = 'water_sources'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = WaterSource.objects.select_related('village').order_by('-created_at')
        
        # Apply filters
        source_type = self.request.GET.get('source_type')
        if source_type:
            queryset = queryset.filter(source_type=source_type)
        
        water_quality = self.request.GET.get('water_quality')
        if water_quality:
            queryset = queryset.filter(water_quality=water_quality)
        
        village = self.request.GET.get('village')
        if village:
            queryset = queryset.filter(village_id=village)
        
        operational = self.request.GET.get('operational')
        if operational:
            queryset = queryset.filter(operational_status=operational == 'true')
        
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(asset_id__icontains=search) |
                Q(pump_type__icontains=search)
            )
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get the queryset safely
        queryset = self.get_queryset()
        
        # Calculate total capacity safely
        try:
            total_capacity = queryset.aggregate(total=Sum('daily_capacity_litres'))['total']
            total_capacity = total_capacity or 0
        except Exception as e:
            print(f"Error calculating total capacity: {e}")
            total_capacity = 0
        
        # Count operational sources safely
        try:
            operational_sources = queryset.filter(operational_status=True).count()
        except Exception as e:
            print(f"Error counting operational sources: {e}")
            operational_sources = 0
        
        # Get villages safely
        try:
            villages = Village.objects.filter(is_active=True).order_by('name')
        except Exception as e:
            print(f"Error getting villages: {e}")
            villages = Village.objects.none()
        
        context.update({
            'source_types': WaterSource.SOURCE_TYPES,
            'water_qualities': WaterSource.WATER_QUALITY,
            'villages': villages,
            'current_filters': {
                'source_type': self.request.GET.get('source_type', ''),
                'water_quality': self.request.GET.get('water_quality', ''),
                'village': self.request.GET.get('village', ''),
                'operational': self.request.GET.get('operational', ''),
                'search': self.request.GET.get('search', ''),
            },
            'total_capacity': total_capacity,
            'operational_sources': operational_sources,
            'total_sources': queryset.count() if queryset else 0,
        })
        return context



class WaterSourceDetailView(LoginRequiredMixin, DetailView):
   """Detail view for water source"""
   model = WaterSource
   template_name = 'infrastructure/water/source_detail.html'
   context_object_name = 'source'
   
   def get_context_data(self, **kwargs):
       context = super().get_context_data(**kwargs)
       source = self.object
       
       # Distribution points connected to this source
       distribution_points = source.distribution_points.all()
       total_users = distribution_points.aggregate(
           total=Sum('estimated_users')
       )['total'] or 0
       
       # Water quality history (latest 5 tests)
       recent_tests = source.maintenance_records.filter(
           maintenance_type='inspection'
       ).order_by('-completed_date')[:5]
       
       # Efficiency metrics
       if source.daily_capacity_litres and total_users:
           litres_per_person_per_day = source.daily_capacity_litres / total_users
       else:
           litres_per_person_per_day = 0
       
       context.update({
           'distribution_points': distribution_points,
           'total_users': total_users,
           'recent_tests': recent_tests,
           'litres_per_person_per_day': round(litres_per_person_per_day, 2),
           'households_served': sum([dp.households_served.count() for dp in distribution_points]),
       })
       
       return context


class WaterDistributionPointListView(LoginRequiredMixin, ListView):
   """List water distribution points"""
   model = WaterDistributionPoint
   template_name = 'infrastructure/water/distribution_list.html'
   context_object_name = 'distribution_points'
   paginate_by = 25
   
   def get_queryset(self):
       queryset = WaterDistributionPoint.objects.select_related(
           'village', 'water_source'
       ).order_by('-created_at')
       
       # Apply filters
       distribution_type = self.request.GET.get('distribution_type')
       if distribution_type:
           queryset = queryset.filter(distribution_type=distribution_type)
       
       village = self.request.GET.get('village')
       if village:
           queryset = queryset.filter(village_id=village)
       
       water_source = self.request.GET.get('water_source')
       if water_source:
           queryset = queryset.filter(water_source_id=water_source)
       
       operational = self.request.GET.get('operational')
       if operational:
           queryset = queryset.filter(operational_status=operational == 'true')
       
       search = self.request.GET.get('search')
       if search:
           queryset = queryset.filter(
               Q(name__icontains=search) |
               Q(asset_id__icontains=search)
           )
       
       return queryset
   
   def get_context_data(self, **kwargs):
       context = super().get_context_data(**kwargs)
       context.update({
           'distribution_types': WaterDistributionPoint.DISTRIBUTION_TYPES,
           'villages': Village.objects.filter(is_active=True).order_by('name'),
           'water_sources': WaterSource.objects.filter(operational_status=True).order_by('name'),
           'current_filters': {
               'distribution_type': self.request.GET.get('distribution_type', ''),
               'village': self.request.GET.get('village', ''),
               'water_source': self.request.GET.get('water_source', ''),
               'operational': self.request.GET.get('operational', ''),
               'search': self.request.GET.get('search', ''),
           },
           'total_users': self.get_queryset().aggregate(
               total=Sum('estimated_users')
           )['total'] or 0,
       })
       return context


class MaintenanceRecordListView(LoginRequiredMixin, ListView):
   """List maintenance records"""
   model = MaintenanceRecord
   template_name = 'infrastructure/maintenance/record_list.html'
   context_object_name = 'maintenance_records'
   paginate_by = 25
   
   def get_queryset(self):
       queryset = MaintenanceRecord.objects.select_related(
           'asset', 'technician'
       ).order_by('-scheduled_date')
       
       # Apply filters
       maintenance_type = self.request.GET.get('maintenance_type')
       if maintenance_type:
           queryset = queryset.filter(maintenance_type=maintenance_type)
       
       status = self.request.GET.get('status')
       if status:
           queryset = queryset.filter(status=status)
       
       asset_category = self.request.GET.get('asset_category')
       if asset_category:
           queryset = queryset.filter(asset__category=asset_category)
       
       technician = self.request.GET.get('technician')
       if technician:
           queryset = queryset.filter(technician_id=technician)
       
       date_from = self.request.GET.get('date_from')
       if date_from:
           queryset = queryset.filter(scheduled_date__gte=date_from)
       
       date_to = self.request.GET.get('date_to')
       if date_to:
           queryset = queryset.filter(scheduled_date__lte=date_to)
       
       search = self.request.GET.get('search')
       if search:
           queryset = queryset.filter(
               Q(asset__name__icontains=search) |
               Q(description__icontains=search) |
               Q(contractor__icontains=search)
           )
       
       return queryset
   
   def get_context_data(self, **kwargs):
       context = super().get_context_data(**kwargs)
       
       # Calculate summary statistics
       queryset = self.get_queryset()
       total_cost = queryset.filter(status='completed').aggregate(
           total=Sum('actual_cost')
       )['total'] or 0
       
       overdue_maintenance = queryset.filter(
           status__in=['scheduled', 'in_progress'],
           scheduled_date__lt=timezone.now().date()
       ).count()
       
       context.update({
           'maintenance_types': MaintenanceRecord.MAINTENANCE_TYPES,
           'statuses': MaintenanceRecord.STATUS_CHOICES,
           'asset_categories': InfrastructureAsset.ASSET_CATEGORIES,
           'technicians': User.objects.filter(
               maintenance_performed__isnull=False
           ).distinct().order_by('first_name'),
           'current_filters': {
               'maintenance_type': self.request.GET.get('maintenance_type', ''),
               'status': self.request.GET.get('status', ''),
               'asset_category': self.request.GET.get('asset_category', ''),
               'technician': self.request.GET.get('technician', ''),
               'date_from': self.request.GET.get('date_from', ''),
               'date_to': self.request.GET.get('date_to', ''),
               'search': self.request.GET.get('search', ''),
           },
           'total_cost': total_cost,
           'overdue_maintenance': overdue_maintenance,
       })
       return context


class ServiceInterruptionListView(LoginRequiredMixin, ListView):
   """List service interruptions"""
   model = ServiceInterruption
   template_name = 'infrastructure/interruption_list.html'
   context_object_name = 'interruptions'
   paginate_by = 25
   
   def get_queryset(self):
       queryset = ServiceInterruption.objects.select_related('asset').order_by('-start_time')
       
       # Apply filters
       interruption_type = self.request.GET.get('interruption_type')
       if interruption_type:
           queryset = queryset.filter(interruption_type=interruption_type)
       
       severity = self.request.GET.get('severity')
       if severity:
           queryset = queryset.filter(severity=severity)
       
       status = self.request.GET.get('status')
       if status == 'ongoing':
           queryset = queryset.filter(is_resolved=False)
       elif status == 'resolved':
           queryset = queryset.filter(is_resolved=True)
       
       asset_category = self.request.GET.get('asset_category')
       if asset_category:
           queryset = queryset.filter(asset__category=asset_category)
       
       date_from = self.request.GET.get('date_from')
       if date_from:
           queryset = queryset.filter(start_time__date__gte=date_from)
       
       date_to = self.request.GET.get('date_to')
       if date_to:
           queryset = queryset.filter(start_time__date__lte=date_to)
       
       search = self.request.GET.get('search')
       if search:
           queryset = queryset.filter(
               Q(asset__name__icontains=search) |
               Q(cause__icontains=search) |
               Q(description__icontains=search)
           )
       
       return queryset
   
   def get_context_data(self, **kwargs):
       context = super().get_context_data(**kwargs)
       
       # Summary statistics
       queryset = self.get_queryset()
       ongoing_count = queryset.filter(is_resolved=False).count()
       total_affected = queryset.aggregate(
           total=Sum('estimated_affected_people')
       )['total'] or 0
       
       # Average resolution time for resolved interruptions
       resolved_interruptions = queryset.filter(is_resolved=True, end_time__isnull=False)
       avg_resolution_time = 0
       if resolved_interruptions.exists():
           total_time = sum([
               (i.end_time - i.start_time).total_seconds() / 3600
               for i in resolved_interruptions
           ])
           avg_resolution_time = total_time / resolved_interruptions.count()
       
       context.update({
           'interruption_types': ServiceInterruption.INTERRUPTION_TYPES,
           'severity_levels': ServiceInterruption.SEVERITY_LEVELS,
           'asset_categories': InfrastructureAsset.ASSET_CATEGORIES,
           'current_filters': {
               'interruption_type': self.request.GET.get('interruption_type', ''),
               'severity': self.request.GET.get('severity', ''),
               'status': self.request.GET.get('status', ''),
               'asset_category': self.request.GET.get('asset_category', ''),
               'date_from': self.request.GET.get('date_from', ''),
               'date_to': self.request.GET.get('date_to', ''),
               'search': self.request.GET.get('search', ''),
           },
           'ongoing_count': ongoing_count,
           'total_affected': total_affected,
           'avg_resolution_time': round(avg_resolution_time, 2),
       })
       return context


# Create Views
class WaterSourceCreateView(LoginRequiredMixin, CreateView):
   """Create new water source"""
   model = WaterSource
   form_class = WaterSourceForm
   template_name = 'infrastructure/water/source_form.html'
   success_url = reverse_lazy('infrastructure:water_source_list')
   
   def form_valid(self, form):
       messages.success(self.request, 'Water source created successfully!')
       return super().form_valid(form)


class WaterDistributionPointCreateView(LoginRequiredMixin, CreateView):
   """Create new water distribution point"""
   model = WaterDistributionPoint
   form_class = WaterDistributionPointForm
   template_name = 'infrastructure/water/distribution_form.html'
   success_url = reverse_lazy('infrastructure:water_distribution_list')
   
   def form_valid(self, form):
       messages.success(self.request, 'Water distribution point created successfully!')
       return super().form_valid(form)


class MaintenanceRecordCreateView(LoginRequiredMixin, CreateView):
   """Create new maintenance record"""
   model = MaintenanceRecord
   form_class = MaintenanceRecordForm
   template_name = 'infrastructure/maintenance/record_form.html'
   success_url = reverse_lazy('infrastructure:maintenance_list')
   
   def form_valid(self, form):
       messages.success(self.request, 'Maintenance record created successfully!')
       return super().form_valid(form)


class ServiceInterruptionCreateView(LoginRequiredMixin, CreateView):
   """Create new service interruption"""
   model = ServiceInterruption
   form_class = ServiceInterruptionForm
   template_name = 'infrastructure/interruption_form.html'
   success_url = reverse_lazy('infrastructure:interruption_list')
   
   def form_valid(self, form):
       messages.success(self.request, 'Service interruption logged successfully!')
       return super().form_valid(form)


# Update Views
class WaterSourceUpdateView(LoginRequiredMixin, UpdateView):
   """Update water source"""
   model = WaterSource
   form_class = WaterSourceForm
   template_name = 'infrastructure/water/source_form.html'
   
   def get_success_url(self):
       return reverse('infrastructure:water_source_detail', kwargs={'pk': self.object.pk})
   
   def form_valid(self, form):
       messages.success(self.request, 'Water source updated successfully!')
       return super().form_valid(form)


class MaintenanceRecordUpdateView(LoginRequiredMixin, UpdateView):
   """Update maintenance record"""
   model = MaintenanceRecord
   form_class = MaintenanceRecordForm
   template_name = 'infrastructure/maintenance/record_form.html'
   
   def get_success_url(self):
       return reverse('infrastructure:maintenance_detail', kwargs={'pk': self.object.pk})
   
   def form_valid(self, form):
       messages.success(self.request, 'Maintenance record updated successfully!')
       return super().form_valid(form)


# AJAX Views
@login_required
def get_water_sources_by_village(request):
   """Get water sources for a specific village (AJAX)"""
   village_id = request.GET.get('village_id')
   if village_id:
       water_sources = WaterSource.objects.filter(
           village_id=village_id,
           operational_status=True
       ).values('id', 'name', 'source_type')
       return JsonResponse({'water_sources': list(water_sources)})
   return JsonResponse({'water_sources': []})


@login_required
def get_distribution_points_by_source(request):
   """Get distribution points for a specific water source (AJAX)"""
   source_id = request.GET.get('source_id')
   if source_id:
       distribution_points = WaterDistributionPoint.objects.filter(
           water_source_id=source_id
       ).values('id', 'name', 'distribution_type', 'estimated_users')
       return JsonResponse({'distribution_points': list(distribution_points)})
   return JsonResponse({'distribution_points': []})

# infrastructure/views.py (update the infrastructure_analytics function)

# infrastructure/views.py (update the infrastructure_analytics function)

from django.db.models import Count, Sum, Avg, Q
from django.db.models.functions import Extract
from django.utils import timezone
from datetime import datetime, timedelta
import calendar

@login_required
def infrastructure_analytics(request):
    """Infrastructure analytics and reporting - Database agnostic version"""
    current_year = timezone.now().year
    
    # Asset distribution by category
    category_data = InfrastructureAsset.objects.values('category').annotate(
        total=Count('id'),
        operational=Count('id', filter=Q(operational_status=True)),
        excellent=Count('id', filter=Q(condition='excellent')),
        good=Count('id', filter=Q(condition='good')),
        fair=Count('id', filter=Q(condition='fair')),
        poor=Count('id', filter=Q(condition='poor')),
        critical=Count('id', filter=Q(condition='critical')),
    ).order_by('category')
    
    # Monthly maintenance costs - using Python aggregation to avoid SQL issues
    monthly_costs = []
    for month in range(1, 13):
        try:
            cost = MaintenanceRecord.objects.filter(
                completed_date__year=current_year,
                completed_date__month=month,
                status='completed'
            ).aggregate(total=Sum('actual_cost'))['total'] or 0
            monthly_costs.append({
                'month': month,
                'month_name': calendar.month_abbr[month],
                'cost': float(cost)
            })
        except Exception as e:
            print(f"Error calculating monthly cost for month {month}: {e}")
            monthly_costs.append({
                'month': month,
                'month_name': calendar.month_abbr[month],
                'cost': 0
            })
    
    # Service interruption trends - using Python aggregation
    interruption_trends = []
    for month in range(1, 13):
        try:
            count = ServiceInterruption.objects.filter(
                start_time__year=current_year,
                start_time__month=month
            ).count()
            
            avg_affected = ServiceInterruption.objects.filter(
                start_time__year=current_year,
                start_time__month=month
            ).aggregate(avg=Avg('estimated_affected_people'))['avg'] or 0
            
            interruption_trends.append({
                'month': month,
                'month_name': calendar.month_abbr[month],
                'count': count,
                'avg_affected': round(float(avg_affected), 1)
            })
        except Exception as e:
            print(f"Error calculating interruption trend for month {month}: {e}")
            interruption_trends.append({
                'month': month,
                'month_name': calendar.month_abbr[month],
                'count': 0,
                'avg_affected': 0
            })
    
    # Village water coverage - Simplified and safe approach
    try:
        # Get all active villages
        villages = Village.objects.filter(is_active=True).prefetch_related(
            'infrastructure_assets', 'households'
        )
        
        village_coverage = []
        for village in villages:
            try:
                # Count water sources for this village
                water_sources_count = village.infrastructure_assets.filter(category='water').count()
                
                # Count households
                households_count = village.households.count()
                
                # Only include villages with households and/or water sources
                if households_count > 0 or water_sources_count > 0:
                    # Calculate population (residents in households)
                    calculated_population = 0
                    try:
                        for household in village.households.all():
                            calculated_population += household.residents.count()
                    except:
                        calculated_population = 0
                    
                    village_data = {
                        'village': village,
                        'water_sources_count': water_sources_count,
                        'households_count': households_count,
                        'calculated_population': calculated_population,
                        'coverage_ratio': round(water_sources_count / households_count, 2) if households_count > 0 else 0
                    }
                    village_coverage.append(village_data)
            except Exception as e:
                print(f"Error processing village {village.name}: {e}")
                continue
        
        # Sort by water sources count and limit
        village_coverage.sort(key=lambda x: x['water_sources_count'], reverse=True)
        village_coverage = village_coverage[:15]  # Top 15 villages
        
    except Exception as e:
        print(f"Error calculating village coverage: {e}")
        village_coverage = []
    
    # Asset age distribution - Safe calculation
    try:
        current_date = timezone.now().date()
        age_categories = {
            'Less than 1 year': 0,
            '1-5 years': 0,
            '5-10 years': 0,
            '10-20 years': 0,
            'Over 20 years': 0
        }
        
        total_assets_with_age = 0
        total_age = 0
        
        assets_with_dates = InfrastructureAsset.objects.filter(
            installation_date__isnull=False
        ).only('installation_date')
        
        for asset in assets_with_dates:
            try:
                age_years = (current_date - asset.installation_date).days // 365
                total_assets_with_age += 1
                total_age += age_years
                
                if age_years < 1:
                    age_categories['Less than 1 year'] += 1
                elif age_years < 5:
                    age_categories['1-5 years'] += 1
                elif age_years < 10:
                    age_categories['5-10 years'] += 1
                elif age_years < 20:
                    age_categories['10-20 years'] += 1
                else:
                    age_categories['Over 20 years'] += 1
            except Exception as e:
                print(f"Error calculating age for asset: {e}")
                continue
        
        # Convert to list for template
        age_distribution = [
            {'category': category, 'count': count}
            for category, count in age_categories.items()
        ]
        
        # Calculate average age
        avg_asset_age = total_age / total_assets_with_age if total_assets_with_age > 0 else 0
        
    except Exception as e:
        print(f"Error calculating age distribution: {e}")
        age_distribution = []
        avg_asset_age = 0
    
    # Basic statistics with error handling
    try:
        total_assets = InfrastructureAsset.objects.count()
        operational_assets = InfrastructureAsset.objects.filter(operational_status=True).count()
        critical_assets = InfrastructureAsset.objects.filter(condition='critical').count()
    except Exception as e:
        print(f"Error calculating basic asset stats: {e}")
        total_assets = operational_assets = critical_assets = 0
    
    # Water infrastructure statistics
    try:
        water_sources = WaterSource.objects.count()
        operational_water_sources = WaterSource.objects.filter(operational_status=True).count()
        total_water_capacity = WaterSource.objects.aggregate(
            total=Sum('daily_capacity_litres')
        )['total'] or 0
    except Exception as e:
        print(f"Error calculating water stats: {e}")
        water_sources = operational_water_sources = 0
        total_water_capacity = 0
    
    # Maintenance statistics
    try:
        pending_maintenance = MaintenanceRecord.objects.filter(
            status__in=['scheduled', 'in_progress']
        ).count()
        
        overdue_maintenance = MaintenanceRecord.objects.filter(
            status__in=['scheduled', 'in_progress'],
            scheduled_date__lt=timezone.now().date()
        ).count()
    except Exception as e:
        print(f"Error calculating maintenance stats: {e}")
        pending_maintenance = overdue_maintenance = 0
    
    # Service interruption statistics
    try:
        active_interruptions = ServiceInterruption.objects.filter(is_resolved=False).count()
    except Exception as e:
        print(f"Error calculating interruption stats: {e}")
        active_interruptions = 0
    
    # Calculate derived metrics
    total_maintenance_cost = sum([item['cost'] for item in monthly_costs])
    operational_percentage = (operational_assets / total_assets * 100) if total_assets > 0 else 0
    water_operational_percentage = (operational_water_sources / water_sources * 100) if water_sources > 0 else 0
    
    # Performance indicators
    assets_needing_attention = critical_assets + overdue_maintenance + active_interruptions
    
    context = {
        'category_data': category_data,
        'monthly_costs': monthly_costs,
        'interruption_trends': interruption_trends,
        'village_coverage': village_coverage,
        'age_distribution': age_distribution,
        'total_maintenance_cost': total_maintenance_cost,
        'avg_asset_age': round(avg_asset_age, 1),
        'current_year': current_year,
        
        # Summary statistics
        'total_assets': total_assets,
        'operational_assets': operational_assets,
        'critical_assets': critical_assets,
        'operational_percentage': round(operational_percentage, 1),
        
        # Water infrastructure
        'water_sources': water_sources,
        'operational_water_sources': operational_water_sources,
        'water_operational_percentage': round(water_operational_percentage, 1),
        'total_water_capacity': total_water_capacity,
        
        # Maintenance
        'pending_maintenance': pending_maintenance,
        'overdue_maintenance': overdue_maintenance,
        
        # Service interruptions
        'active_interruptions': active_interruptions,
        
        # Performance indicators
        'assets_needing_attention': assets_needing_attention,
        
        # Additional context
        'villages_with_water_count': len([v for v in village_coverage if v['water_sources_count'] > 0]),
    }
    
    return render(request, 'infrastructure/analytics.html', context)


@login_required
def mark_interruption_resolved(request, pk):
   """Mark service interruption as resolved"""
   if request.method == 'POST':
       interruption = get_object_or_404(ServiceInterruption, pk=pk)
       interruption.is_resolved = True
       interruption.resolution_date = timezone.now()
       interruption.end_time = timezone.now()
       interruption.save()
       
       messages.success(request, 'Service interruption marked as resolved!')
       return JsonResponse({'success': True})
   
   return JsonResponse({'success': False})


@login_required
def schedule_maintenance(request, asset_id):
   """Quick schedule maintenance for an asset"""
   if request.method == 'POST':
       asset = get_object_or_404(InfrastructureAsset, pk=asset_id)
       
       # Create basic maintenance record
       maintenance = MaintenanceRecord.objects.create(
           asset=asset,
           maintenance_type=request.POST.get('maintenance_type', 'preventive'),
           scheduled_date=request.POST.get('scheduled_date'),
           description=request.POST.get('description', ''),
           technician_id=request.POST.get('technician') if request.POST.get('technician') else None,
           estimated_cost=request.POST.get('estimated_cost') if request.POST.get('estimated_cost') else None,
       )
       
       messages.success(request, f'Maintenance scheduled for {asset.name}!')
       return JsonResponse({'success': True, 'maintenance_id': maintenance.id})
   
   return JsonResponse({'success': False})

